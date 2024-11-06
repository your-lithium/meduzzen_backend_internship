from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import Notification, Quiz, StatusEnum
from app.db.repo.notification import NotificationRepo
from app.db.repo.quiz import QuizRepo
from tests import payload
from tests.conftest import assert_real_matches_expected, get_user_and_company_ids


@pytest.mark.asyncio
@pytest.mark.parametrize("fill_db_with_memberships", [StatusEnum.MEMBER], indirect=True)
async def test_create_quiz(
    fill_db_with_quizzes,
    fill_db_with_memberships,
    client: AsyncClient,
    test_session: AsyncSession,
):
    _, company_id = await get_user_and_company_ids(
        company_name=payload.test_company_1.name, session=test_session
    )
    response = await client.post(
        f"/quizzes/{company_id}", json=payload.test_quiz_3.model_dump()
    )
    assert response.status_code == 200
    quiz = response.json()
    expected_quiz = {
        **payload.expected_test_quiz_3,
        "company_id": company_id,
    }
    assert_real_matches_expected(quiz, expected_quiz)

    notifications = await NotificationRepo.get_all_by_fields(
        fields=[Notification.text],
        values=[
            (
                f"There's a new quiz {quiz['id']} created by company {company_id}. "
                "You should take it!"
            )
        ],
        session=test_session,
    )
    assert notifications != []


@pytest.mark.asyncio
async def test_get_quizzes_by_company(
    fill_db_with_quizzes, client: AsyncClient, test_session: AsyncSession
):
    _, company_id = await get_user_and_company_ids(
        company_name=payload.test_company_1.name, session=test_session
    )
    response = await client.get(f"/quizzes/{company_id}")
    assert response.status_code == 200

    quizzes = response.json()
    assert quizzes != []

    expected_quizzes = [
        payload.expected_test_quiz_1,
        payload.expected_test_quiz_2,
        payload.expected_test_quiz_2,
    ]
    company_names = [
        payload.test_company_1.name,
        payload.test_company_1.name,
        payload.test_company_2.name,
    ]
    for quiz, expected_quiz, company_name in zip(
        quizzes, expected_quizzes, company_names
    ):
        expected_quiz = {
            **expected_quiz,
            "company_id": (
                await get_user_and_company_ids(
                    company_name=company_name, session=test_session
                )
            )[1],
        }
        assert_real_matches_expected(quiz, expected_quiz)


@pytest.mark.asyncio
async def test_update_quiz(
    fill_db_with_quizzes, client: AsyncClient, test_session: AsyncSession
):
    _, company_id = await get_user_and_company_ids(
        company_name=payload.test_company_1.name, session=test_session
    )
    quiz = await QuizRepo.get_by_fields(
        fields=[Quiz.name, Quiz.company_id],
        values=[payload.test_quiz_1.name, company_id],
        session=test_session,
    )
    assert quiz is not None, "Quiz not found"
    quiz_id = quiz.id
    response = await client.patch(
        f"/quizzes/{quiz_id}", json=payload.test_quiz_1_update.model_dump()
    )
    assert response.status_code == 200
    updated_quiz = response.json()
    expected_quiz = {
        **payload.expected_test_quiz_1_update,
        "company_id": company_id,
    }
    assert_real_matches_expected(updated_quiz, expected_quiz)


@pytest.mark.asyncio
async def test_delete_quiz(
    fill_db_with_quizzes, client: AsyncClient, test_session: AsyncSession
):
    quiz = await QuizRepo.get_by_fields(
        fields=[Quiz.name], values=[payload.test_quiz_1.name], session=test_session
    )
    assert quiz is not None, "Quiz not found"
    quiz_id = quiz.id
    await client.delete(f"/quizzes/{quiz_id}")
    quiz = await QuizRepo.get_by_id(record_id=quiz_id, session=test_session)
    assert quiz is None


@pytest.mark.parametrize("fill_db_with_memberships", [StatusEnum.MEMBER], indirect=True)
@pytest.mark.asyncio
async def test_import_create_quiz(
    fill_db_with_memberships,
    fill_db_with_companies,
    client: AsyncClient,
    test_session: AsyncSession,
):
    _, company_id = await get_user_and_company_ids(
        company_name=payload.test_company_1.name, session=test_session
    )
    with open("tests/payload_files/quiz_create.xlsx", "rb") as file:
        files = {
            "quiz_table": (
                "file.xlsx",
                file,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = await client.put(f"/quizzes/{company_id}/import", files=files)

    assert response.status_code == 200
    quiz = response.json()
    expected_quiz = {
        **payload.expected_test_quiz_1,
        "company_id": company_id,
    }
    assert_real_matches_expected(quiz, expected_quiz)

    notifications = await NotificationRepo.get_all_by_fields(
        fields=[Notification.text],
        values=[
            (
                f"There's a new quiz {quiz['id']} created by company {company_id}. "
                "You should take it!"
            )
        ],
        session=test_session,
    )
    assert notifications != []


@pytest.mark.asyncio
async def test_import_update_quiz(
    fill_db_with_quizzes, client: AsyncClient, test_session: AsyncSession
):
    _, company_id = await get_user_and_company_ids(
        company_name=payload.test_company_1.name, session=test_session
    )
    quiz = await QuizRepo.get_by_fields(
        fields=[Quiz.name, Quiz.company_id],
        values=[payload.test_quiz_1.name, company_id],
        session=test_session,
    )
    assert quiz is not None, "Quiz not found"
    quiz_id = quiz.id

    file_path = "tests/payload_files/quiz_update.xlsx"
    workbook = load_workbook(filename=file_path)
    sheet = workbook["Quiz"]

    sheet.cell(row=1, column=2).value = str(quiz_id)

    with BytesIO() as file_stream:
        workbook.save(file_stream)
        file_stream.seek(0)
        files = {
            "quiz_table": (
                "file.xlsx",
                file_stream,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = await client.put(f"/quizzes/{company_id}/import", files=files)

    assert response.status_code == 200
    updated_quiz = response.json()
    expected_quiz = {
        **payload.expected_test_quiz_1_update,
        "company_id": company_id,
    }
    assert_real_matches_expected(updated_quiz, expected_quiz)
