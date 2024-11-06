import re
from io import BytesIO
from uuid import UUID

from fastapi import Depends, UploadFile
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.database import get_session
from app.db.models import Company, Membership, Quiz, User
from app.db.repo.quiz import QuizRepo
from app.schemas.membership_schemas import MembershipActionRequest
from app.schemas.quiz_schemas import (
    Answer,
    Question,
    QuestionList,
    QuizCreateRequest,
    QuizUpdateRequest,
)
from app.services.company import CompanyService, get_company_service
from app.services.exceptions import (
    MembershipNotFoundError,
    QuizNotFoundError,
    UnsupportedFileFormatError,
)
from app.services.membership import MembershipService, get_membership_service
from app.services.notification import NotificationService, get_notification_service
from app.services.permissions import PermissionService
from app.services.user import UserService, get_user_service


def get_quiz_service(
    user_service=Depends(get_user_service),
    company_service=Depends(get_company_service),
    membership_service=Depends(get_membership_service),
    notification_service=Depends(get_notification_service),
):
    return QuizService(
        user_service, company_service, membership_service, notification_service
    )


class QuizService:
    """Represents a service for handling requests to Quiz model."""

    def __init__(
        self,
        user_service: UserService,
        company_service: CompanyService,
        membership_service: MembershipService,
        notification_service: NotificationService,
    ) -> None:
        self._user_service = user_service
        self._company_service = company_service
        self._membership_service = membership_service
        self._notification_service = notification_service

    async def check_company_and_user(
        self,
        company_id: UUID,
        current_user: User,
        session: AsyncSession = Depends(get_session),
    ) -> None:
        """Check if a Company exists and if a User is its owner or admin.

        Args:
            company_id (UUID): The ID of a Company to check.
            current_user (User): The User to check.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.
        """
        existing_company = await self._company_service.get_company_by_id(
            company_id=company_id, current_user=current_user, session=session
        )

        parties = MembershipActionRequest(
            company_id=company_id, user_id=current_user.id
        )
        membership: Membership | None = None
        try:
            membership = await self._membership_service.get_membership_by_parties(
                parties=parties, session=session
            )
        except MembershipNotFoundError:
            pass
        finally:
            PermissionService.grant_owner_admin_permission(
                owner_id=existing_company.owner_id,
                membership=membership,
                current_user_id=current_user.id,
                operation="update or view a quiz",
            )

    async def get_quizzes_by_company(
        self,
        company_id: UUID,
        limit: int | None = 10,
        offset: int = 0,
        session: AsyncSession = Depends(get_session),
    ) -> list[Quiz]:
        """Get a list of quizzes belonging to one Company.

        Args:
            company_id (UUID): The ID of the company to check.
            limit (int | None, optional):
                How much quizzes to get. Defaults to 10.
                If None, retrieve all records.
            offset (int, optional):
                Where to start getting quizzes. Defaults to 0.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.

        Returns:
            list[Quiz]: The list of quizzes.
        """
        quizzes: list[Quiz] = await QuizRepo.get_quizzes_by_company(
            company_id=company_id,
            limit=limit,
            offset=offset,
            session=session,
        )
        return quizzes

    async def get_quiz_by_id(
        self,
        quiz_id: UUID,
        session: AsyncSession = Depends(get_session),
    ) -> Quiz:
        """Get details for one quiz via its ID.

        Args:
            quiz_id (UUID): The quiz's ID.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.

        Raises:
            QuizNotFoundError: If there's no Quiz with given ID.

        Returns:
            Quiz: Quiz details.
        """
        quiz: Quiz | None = await QuizRepo.get_by_id(record_id=quiz_id, session=session)

        if quiz is None:
            raise QuizNotFoundError(quiz_id)

        return quiz

    async def notify_members_of_created_quiz(
        self,
        new_quiz_id: UUID,
        company_id: UUID,
        session: AsyncSession = Depends(get_session),
    ) -> None:
        """Send notifications to all members of the Company about the new Quiz
        they should take.

        Args:
            new_quiz_id (UUID): The ID of the new Quiz.
            company_id (UUID): The ID of the Company.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.
        """
        members = await self._membership_service.get_members_by_company(
            company_id=company_id, limit=None, offset=0, session=session
        )
        await self._notification_service.bulk_send_notifications(
            users=members,
            text=str(
                f"There's a new quiz {new_quiz_id} created "
                f"by company {company_id}. You should take it!",
            ),
            session=session,
        )

    async def create_quiz(
        self,
        quiz: QuizCreateRequest,
        company_id: UUID,
        current_user: User,
        session: AsyncSession = Depends(get_session),
    ) -> Quiz:
        """Create a new quiz.

        Args:
            quiz (QuizCreateRequest): Details for creating a new quiz.
            company_id (UUID): The company which the quiz should belong to.
            current_user (User): The current user to authorize as an owner or admin.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.

        Returns:
            Quiz: Details of the new quiz.
        """
        await self.check_company_and_user(
            company_id=company_id,
            current_user=current_user,
            session=session,
        )

        new_quiz: Quiz = await QuizRepo.create_quiz(
            quiz=quiz, company_id=company_id, session=session
        )

        await self.notify_members_of_created_quiz(
            new_quiz_id=new_quiz.id, company_id=company_id, session=session
        )
        return new_quiz

    async def update_quiz(
        self,
        quiz_id: UUID,
        quiz_update: QuizUpdateRequest,
        current_user: User,
        session: AsyncSession = Depends(get_session),
    ) -> Company:
        """Update an existing quiz.

        Args:
            quiz_id (UUID): The ID of the quiz to update.
            quiz_update (QuizUpdateRequest): The details which to update in a quiz.
            current_user (User): The current user to authorize as an owner or admin.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.

        Returns:
            Company: Details of the updated quiz.
        """
        existing_quiz = await self.get_quiz_by_id(quiz_id=quiz_id, session=session)

        await self.check_company_and_user(
            company_id=existing_quiz.company_id,
            current_user=current_user,
            session=session,
        )

        updated_quiz: Quiz = await QuizRepo.update_quiz(
            existing_quiz=existing_quiz,
            quiz_update=quiz_update,
            session=session,
        )

        return updated_quiz

    async def delete_quiz(
        self,
        quiz_id: UUID,
        current_user: User,
        session: AsyncSession = Depends(get_session),
    ) -> None:
        """Delete a quiz.

        Args:
            quiz_id (UUID): The ID of the quiz to delete.
            current_user (User): The current user to authorize as an owner or admin.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.
        """
        quiz: Quiz = await self.get_quiz_by_id(quiz_id=quiz_id, session=session)

        await self.check_company_and_user(
            company_id=quiz.company_id,
            current_user=current_user,
            session=session,
        )

        await QuizRepo.delete(entity=quiz, session=session)

    async def extract_answers_from_import(
        self, question_column: list[str]
    ) -> tuple[dict[int, str], list[int]]:
        """Parse the answer options and correct options from the user data.
        Splits by semicolons, strips of whitespaces, transforms the human-readable
        format into the format for Answer schema.

        Args:
            question_column (list[str]): The question column with data.

        Returns:
            tuple[dict[int, str], list[int]]: The extracted options and correct options.
        """
        options = {
            index: text.strip()
            for index, text in enumerate(question_column[2].split(";"))
        }
        correct = [
            index
            for index, text in options.items()
            if text in re.split(r";\s*", question_column[3])
        ]
        return options, correct

    async def extract_full_questions_from_import(
        self, questions_range: list[list[str]]
    ) -> list[Question]:
        """Extracts the full questions from imported data.
        Implies all questions are supposed to be parsed and added.
        Used for quiz creation only.

        Args:
            questions_range (list[list[str]]):
                The questions range from the imported data.

        Returns:
            list[Question]: The extracted questions.
        """
        questions = []
        for question_column in questions_range:
            options, correct = await self.extract_answers_from_import(
                question_column=question_column
            )
            questions.append(
                Question(
                    question=question_column[1],
                    answers=[Answer(options=options, correct=correct)],
                )
            )
        return questions

    async def extract_and_compare_questions_from_import(
        self, questions_range: list[list[str]], existing_questions: QuestionList
    ) -> list[Question]:
        """Extracts and compares the questions from the imported data with the original
        ones. Implies there may be questions the user wants to leave as they are,
        ones they want to delete, and ones they want to update.
        Used for quiz updates only.

        Args:
            questions_range (list[list[str]]):
                The questions range from the imported data.
            existing_questions (QuestionList): The questions from the original quiz.

        Returns:
            list[Question]: The extracted questions.
        """
        questions = []
        for i, question_column in enumerate(questions_range):
            if not question_column[0]:
                continue
            elif any(item is None for item in question_column[1:3]):
                questions.append(existing_questions[i])
            else:
                options, correct = await self.extract_answers_from_import(
                    question_column=question_column
                )
                questions.append(
                    Question(
                        question=question_column[1],
                        answers=[Answer(options=options, correct=correct)],
                    )
                )
        return questions

    async def create_quiz_from_import(
        self,
        info_range: list[str],
        questions_range: list[list[str]],
        company_id: UUID,
        session: AsyncSession = Depends(get_session),
    ) -> Quiz:
        """Create a new Quiz from the imported data.

        Args:
            info_range (list[str]): The main info range from the imported data.
            questions_range (list[list[str]]):
                The questions range from the imported data.
            company_id (UUID): The Company to add the Quiz for.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.

        Returns:
            Quiz: The newly created Quiz.
        """
        questions = await self.extract_full_questions_from_import(
            questions_range=questions_range
        )
        quiz = QuizCreateRequest(
            name=info_range[1],
            description=info_range[2],
            frequency=info_range[3],
            questions=questions,
        )

        new_quiz: Quiz = await QuizRepo.create_quiz(
            quiz=quiz, company_id=company_id, session=session
        )
        await self.notify_members_of_created_quiz(
            new_quiz_id=new_quiz.id, company_id=company_id, session=session
        )
        return new_quiz

    async def update_quiz_from_import(
        self,
        info_range: list[str],
        questions_range: list[list[str]],
        existing_quiz: Quiz,
        session: AsyncSession = Depends(get_session),
    ) -> Quiz:
        """Update an existing Quiz with new info from the imported data.

        Args:
            info_range (list[str]): The main info range from the imported data.
            questions_range (list[list[str]]):
                The questions range from the imported data.
            existing_quiz (Quiz): The existing Quiz to be updated.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.

        Returns:
            Quiz: The newly updated Quiz.
        """
        questions = await self.extract_and_compare_questions_from_import(
            questions_range=questions_range, existing_questions=existing_quiz.questions
        )
        quiz_update = QuizUpdateRequest(
            name=info_range[1] if info_range[1] else None,
            description=info_range[2] if info_range[2] else None,
            frequency=info_range[3] if info_range[3] else None,
            questions=questions,
        )

        updated_quiz: Quiz = await QuizRepo.update_quiz(
            existing_quiz=existing_quiz,
            quiz_update=quiz_update,
            session=session,
        )
        return updated_quiz

    async def import_quiz(
        self,
        company_id: UUID,
        quiz_table: UploadFile,
        current_user: User,
        session: AsyncSession = Depends(get_session),
    ) -> Quiz:
        """Import a Quiz from an Excel file.
        Checks if a new Quiz is supposed to be created or and existing one — updated.

        Args:
            company_id (UUID): The Company to update the Quiz for.
            quiz_table (UploadFile): The file with data to import.
            current_user (User): The current user to authorize as an owner or admin.
            session (AsyncSession):
                The database session used for querying.
                Defaults to the session obtained through get_session.

        Raises:
            UnsupportedFileFormatError:
                If the imported file is not supported by openpyxl.load_workbook().

        Returns:
            Quiz: The resulting created or updated quiz.
        """
        await self.check_company_and_user(
            company_id=company_id,
            current_user=current_user,
            session=session,
        )

        file_contents = await quiz_table.read()
        try:
            workbook = load_workbook(filename=BytesIO(file_contents), data_only=True)
            quiz_sheet = workbook["Quiz"]
        except ValueError:
            raise UnsupportedFileFormatError

        info_range = [quiz_sheet.cell(row=i, column=2).value for i in range(1, 5)]
        questions_range = [
            [quiz_sheet.cell(row=i, column=j).value for i in range(1, 5)]
            for j in range(4, quiz_sheet.max_column + 1)
        ]

        quiz_id = info_range[0]
        if not quiz_id:
            new_quiz = await self.create_quiz_from_import(
                info_range=info_range,
                questions_range=questions_range,
                company_id=company_id,
                session=session,
            )
            return new_quiz
        else:
            existing_quiz = await self.get_quiz_by_id(quiz_id=quiz_id, session=session)
            updated_quiz = await self.update_quiz_from_import(
                info_range=info_range,
                questions_range=questions_range,
                existing_quiz=existing_quiz,
                session=session,
            )
            return updated_quiz
